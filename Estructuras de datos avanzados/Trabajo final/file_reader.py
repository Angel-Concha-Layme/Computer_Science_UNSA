print("Preprocesando datos...")
# read  excel file
from openpyxl import load_workbook
from openpyxl import Workbook

workbook_read = load_workbook(filename="data.xlsx")
sheet_read = workbook_read.active

filename = "processed_data.xlsx"
workbook_write = Workbook()
sheet_write = workbook_write.active

################################################################
# Columnas:

# A: Patient ID (no se considera en la funcion d):      str
# C: SARS-Cov-2 exam result :                           bool (negative | positive)
# B: Patient age quantile:                              int
# G: Hematocrit:                                        float
# I: Platelets:                                         float
# J: Mean platelet volume:                              float
# M: Mean corpuscular hemoglobin concentration (MCHC):  float
# N: Leukocytes:                                        float
# O: Basophils:                                         float
# Q: Eosinophils:                                       float
# S: Monocytes:                                         float
# AP: Proteina C reativa mg/dL:                         float

#  Se cuentan los registro i, para i: 2, 3, 4, ..., 5644, 5645

################################################################


sheet_write["A1"] = "Patient ID"
sheet_write["B1"] = "SARS-Cov-2 exam result"
sheet_write["C1"] = "Patient age quantile"
sheet_write["D1"] = "Hematocrit"
sheet_write["E1"] = "Platelets"
sheet_write["F1"] = "Mean platelet volume"
sheet_write["G1"] = "Mean corpuscular hemoglobin concentration (MCHC)"
sheet_write["H1"] = "Leukocytes"
sheet_write["I1"] = "Basophils"
sheet_write["J1"] = "Eosinophils"
sheet_write["K1"] = "Monocytes"
sheet_write["L1"] = "Proteina C reativa mg/dL"

# get min and max values

# read until end of file
for k in range(2, 5646):
    sheet_write["A" + str(k)] = sheet_read["A" + str(k)].value
    if (sheet_read["C" + str(k)].value == "negative"):
        sheet_write["B" + str(k)] = 0
    elif (sheet_read["C" + str(k)].value == "positive"):
        sheet_write["B" + str(k)] = 1

    if sheet_read["B" + str(k)].value is not None:
        sheet_write["C" + str(k)] = round((float(sheet_read["B" + str(k)].value)+0)/19, 4)
    else:
        sheet_write["C" + str(k)] = -1.0

    if sheet_read["G" + str(k)].value is not None:
        sheet_write["D" + str(k)] = round((float(sheet_read["G" + str(k)].value)+4.0665)/6.5003, 4)
    else:
        sheet_write["D" + str(k)] = -1.0

    if sheet_read["I" + str(k)].value is not None:
        sheet_write["E" + str(k)] = round((float(sheet_read["I" + str(k)].value)+2.276)/5.2761, 4)
    else:
        sheet_write["E" + str(k)] = -1.0
        
    if sheet_read["J" + str(k)].value is not None:
        sheet_write["F" + str(k)] = round((float(sheet_read["J" + str(k)].value)+2.4576)/6.1706, 4)
    else:
        sheet_write["F" + str(k)] = -1.0

    if sheet_read["M" + str(k)].value is not None:
        sheet_write["G" + str(k)] = round((float(sheet_read["M" + str(k)].value)+3.6394)/6.9704, 4)
    else:
        sheet_write["G" + str(k)] = -1.0

    if sheet_read["N" + str(k)].value is not None:
        sheet_write["H" + str(k)] = round((float(sheet_read["N" + str(k)].value)+1.8283)/6.3503, 4)
    else:
        sheet_write["H" + str(k)] = -1.0

    if sheet_read["O" + str(k)].value is not None:
        sheet_write["I" + str(k)] = round((float(sheet_read["O" + str(k)].value)+1.1401)/5.8037, 4)
    else:
        sheet_write["I" + str(k)] = -1.0

    if sheet_read["Q" + str(k)].value is not None:
        sheet_write["J" + str(k)] = round((float(sheet_read["Q" + str(k)].value)+0.8355)/9.1864, 4)
    else:
        sheet_write["J" + str(k)] = -1.0
        
    if sheet_read["S" + str(k)].value is not None:
        sheet_write["K" + str(k)] = round((float(sheet_read["S" + str(k)].value)+2.1637)/6.6709, 4)
    else:
        sheet_write["K" + str(k)] = -1.0

    if sheet_read["AP" + str(k)].value is not None:
        sheet_write["L" + str(k)] = round((float(sheet_read["AP" + str(k)].value)+0.5354)/8.562, 4)
    else:
        sheet_write["L" + str(k)] = -1.0

workbook_write.save(filename=filename)